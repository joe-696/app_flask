from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timezone
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import base64

# Crear la aplicación con las carpetas correctas
app = Flask(__name__, template_folder='app/templates', static_folder='app/static')

# Configuración de la base de datos SQLite
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'restaurante.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'tu_clave_secreta_muy_segura_aqui_2025'

# Inicializar SQLAlchemy
db = SQLAlchemy(app)

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))

# ===== DECORADORES DE PERMISOS =====

def requiere_permiso(rol_requerido):
    """Decorador para verificar permisos por rol"""
    def decorador(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.rol != rol_requerido and current_user.rol != 'admin':
                flash('No tienes permisos para acceder a esta sección', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorador

# ===== MODELOS =====

class Usuario(UserMixin, db.Model):
    """Modelo para usuarios del sistema con roles"""
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=False)
    rol = db.Column(db.String(20), nullable=False, default='mesero')  # admin, mesero, cocinero
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    ultimo_acceso = db.Column(db.DateTime)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def __repr__(self):
        return f'<Usuario {self.username}>'

class Mesa(db.Model):
    """Modelo para las mesas del restaurante"""
    __tablename__ = 'mesas'
    
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.String(10), unique=True, nullable=False)
    capacidad = db.Column(db.Integer, nullable=False, default=4)
    estado = db.Column(db.String(20), default='disponible')  # disponible, ocupada, reservada
    ubicacion = db.Column(db.String(50))  # interior, terraza, vip
    activa = db.Column(db.Boolean, default=True)
    
    # Relación con pedidos
    pedidos = db.relationship('Pedido', backref='mesa_info', lazy=True)
    
    def __repr__(self):
        return f'<Mesa {self.numero}>'

class Producto(db.Model):
    """Modelo para los productos/platillos del restaurante"""
    __tablename__ = 'productos'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(db.Float, nullable=False)
    categoria = db.Column(db.String(50), nullable=False)
    disponible = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    
    # Relación con detalles de pedido
    detalles_pedido = db.relationship('DetallePedido', backref='producto', lazy=True)
    
    def __repr__(self):
        return f'<Producto {self.nombre}>'
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'precio': self.precio,
            'categoria': self.categoria,
            'disponible': self.disponible
        }

class Pedido(db.Model):
    """Modelo para los pedidos del restaurante"""
    __tablename__ = 'pedidos'
    
    id = db.Column(db.Integer, primary_key=True)
    cliente_nombre = db.Column(db.String(100), nullable=False)
    cliente_telefono = db.Column(db.String(20))
    mesa_id = db.Column(db.Integer, db.ForeignKey('mesas.id'))  # Relación con Mesa
    mesa_numero = db.Column(db.String(10))  # Mantener compatibilidad
    estado = db.Column(db.String(20), default='pendiente')
    total = db.Column(db.Float, default=0.0)
    fecha = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    observaciones = db.Column(db.Text)
    
    # Referencias de usuario
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))  # Quien tomó el pedido
    cocinero_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))  # Quien lo preparó
    
    # Relaciones
    detalles = db.relationship('DetallePedido', backref='pedido', lazy=True, cascade='all, delete-orphan')
    usuario = db.relationship('Usuario', foreign_keys=[usuario_id], backref='pedidos_tomados')
    cocinero = db.relationship('Usuario', foreign_keys=[cocinero_id], backref='pedidos_preparados')
    
    # Mantener compatibilidad con propiedad mesa
    @property
    def mesa(self):
        return self.mesa_numero or (self.mesa_info.numero if self.mesa_info else None)
    
    @mesa.setter
    def mesa(self, value):
        self.mesa_numero = value
    
    def __repr__(self):
        return f'<Pedido {self.id} - {self.cliente_nombre}>'
    
    def calcular_total(self):
        """Calcular el total del pedido basado en sus detalles"""
        total = sum(detalle.subtotal for detalle in self.detalles)
        self.total = total
        return total

class DetallePedido(db.Model):
    """Modelo para los detalles de cada pedido"""
    __tablename__ = 'detalles_pedido'
    
    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(db.Integer, db.ForeignKey('pedidos.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('productos.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    precio_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    observaciones = db.Column(db.String(200))
    
    def __repr__(self):
        return f'<DetallePedido {self.producto.nombre} x{self.cantidad}>'
    
    def calcular_subtotal(self):
        """Calcular el subtotal del detalle"""
        self.subtotal = self.cantidad * self.precio_unitario
        return self.subtotal

# ===== RUTAS DE AUTENTICACIÓN =====

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        usuario = Usuario.query.filter_by(username=username).first()
        
        if usuario and usuario.check_password(password) and usuario.activo:
            login_user(usuario)
            usuario.ultimo_acceso = datetime.now(timezone.utc)
            db.session.commit()
            
            next_page = request.args.get('next')
            flash(f'¡Bienvenido {usuario.nombre_completo}!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('Has cerrado sesión exitosamente', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@login_required
@requiere_permiso('admin')
def register():
    """Registro de nuevos usuarios (solo admin)"""
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        nombre_completo = request.form.get('nombre_completo')
        rol = request.form.get('rol')
        
        # Verificar si el usuario ya existe
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            return render_template('auth/register.html')
        
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado', 'error')
            return render_template('auth/register.html')
        
        # Crear nuevo usuario
        usuario = Usuario(
            username=username,
            email=email,
            nombre_completo=nombre_completo,
            rol=rol
        )
        usuario.set_password(password)
        
        db.session.add(usuario)
        db.session.commit()
        
        flash(f'Usuario {username} creado exitosamente', 'success')
        return redirect(url_for('empleados'))
    
    return render_template('auth/register.html')

# ===== RUTAS DE EMPLEADOS =====

@app.route('/empleados')
@login_required
@requiere_permiso('admin')
def empleados():
    """Gestión de empleados"""
    # Filtros
    rol_filtro = request.args.get('rol')
    buscar = request.args.get('buscar')
    
    # Query base
    query = Usuario.query
    
    # Aplicar filtros
    if rol_filtro:
        query = query.filter_by(rol=rol_filtro)
    if buscar:
        query = query.filter(
            (Usuario.nombre_completo.contains(buscar)) | 
            (Usuario.email.contains(buscar)) | 
            (Usuario.username.contains(buscar))
        )
    
    usuarios = query.order_by(Usuario.fecha_creacion.desc()).all()
    
    # Estadísticas
    total_usuarios = Usuario.query.count()
    administradores = Usuario.query.filter_by(rol='admin').count()
    meseros = Usuario.query.filter_by(rol='mesero').count()
    cocineros = Usuario.query.filter_by(rol='cocinero').count()
    
    stats = {
        'total_usuarios': total_usuarios,
        'administradores': administradores,
        'meseros': meseros,
        'cocineros': cocineros
    }
    
    return render_template('empleados/index.html', usuarios=usuarios, stats=stats)

@app.route('/empleados/<int:id>/detalle')
@login_required
@requiere_permiso('admin')
def detalle_empleado(id):
    """Obtener detalles de empleado para modal"""
    try:
        empleado = Usuario.query.get_or_404(id)
        return jsonify({
            'username': empleado.username,
            'email': empleado.email,
            'nombre_completo': empleado.nombre_completo,
            'rol': empleado.rol,
            'activo': empleado.activo,
            'ultimo_acceso': empleado.ultimo_acceso.strftime('%d/%m/%Y %H:%M') if empleado.ultimo_acceso else 'Nunca'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/empleados/<int:id>/toggle-estado', methods=['POST'])
@login_required
@requiere_permiso('admin')
def toggle_estado_empleado(id):
    """Cambiar estado activo/inactivo del empleado"""
    try:
        empleado = Usuario.query.get_or_404(id)
        data = request.get_json()
        empleado.activo = data['activo']
        db.session.commit()
        
        estado = 'activado' if empleado.activo else 'desactivado'
        return jsonify({'success': True, 'message': f'Empleado {empleado.username} {estado}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/empleados/<int:id>/eliminar', methods=['DELETE'])
@login_required
@requiere_permiso('admin')
def eliminar_empleado(id):
    """Eliminar empleado"""
    try:
        empleado = Usuario.query.get_or_404(id)
        
        # No permitir eliminar al usuario actual
        if empleado.id == current_user.id:
            return jsonify({'success': False, 'message': 'No puedes eliminar tu propia cuenta'})
        
        # Verificar si tiene pedidos asociados
        if empleado.pedidos_tomados or empleado.pedidos_preparados:
            return jsonify({'success': False, 'message': 'No se puede eliminar un empleado con pedidos asociados'})
        
        db.session.delete(empleado)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Empleado eliminado exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/empleados/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@requiere_permiso('admin')
def editar_empleado(id):
    """Editar empleado"""
    empleado = Usuario.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            empleado.nombre_completo = request.form.get('nombre_completo')
            empleado.email = request.form.get('email')
            empleado.rol = request.form.get('rol')
            
            # Cambiar contraseña si se proporciona
            new_password = request.form.get('password')
            if new_password:
                empleado.set_password(new_password)
            
            db.session.commit()
            flash(f'Empleado {empleado.username} actualizado exitosamente', 'success')
            return redirect(url_for('empleados'))
        except Exception as e:
            flash(f'Error al actualizar empleado: {str(e)}', 'error')
    
    return render_template('empleados/editar.html', empleado=empleado)

# ===== RUTAS DE MESAS =====

@app.route('/mesas')
@login_required
def mesas():
    """Gestión de mesas (no disponible para cocineros)"""
    # Los cocineros no tienen acceso a gestión de mesas
    if current_user.rol == 'cocinero':
        flash('No tienes permisos para acceder a la gestión de mesas', 'error')
        return redirect(url_for('index'))
        
    # Filtros
    estado_filtro = request.args.get('estado')
    capacidad_filtro = request.args.get('capacidad')
    
    # Query base
    query = Mesa.query.filter_by(activa=True)
    
    # Aplicar filtros
    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)
    if capacidad_filtro:
        query = query.filter_by(capacidad=int(capacidad_filtro))
    
    mesas = query.order_by(Mesa.numero).all()
    
    # Estadísticas
    total_mesas = Mesa.query.filter_by(activa=True).count()
    disponibles = Mesa.query.filter_by(estado='disponible', activa=True).count()
    ocupadas = Mesa.query.filter_by(estado='ocupada', activa=True).count()
    reservadas = Mesa.query.filter_by(estado='reservada', activa=True).count()
    
    # Agregar información de pedidos actuales para mesas ocupadas
    for mesa in mesas:
        if mesa.estado == 'ocupada' and mesa.pedidos:
            # Buscar el pedido activo más reciente (no entregado ni cancelado)
            pedido_actual = None
            for pedido in reversed(mesa.pedidos):  # Empezar por el más reciente
                if pedido.estado not in ['entregado', 'cancelado']:
                    pedido_actual = pedido
                    break
            
            if pedido_actual:
                mesa.pedido_actual = pedido_actual
                # Asegurar que ambos datetimes tengan la misma zona horaria
                ahora = datetime.now(timezone.utc)
                fecha_pedido = pedido_actual.fecha
                if fecha_pedido.tzinfo is None:
                    fecha_pedido = fecha_pedido.replace(tzinfo=timezone.utc)
                
                tiempo_ocupada = ahora - fecha_pedido
                horas = int(tiempo_ocupada.total_seconds() // 3600)
                minutos = int((tiempo_ocupada.total_seconds() % 3600) // 60)
                mesa.tiempo_ocupada = f"{horas}h {minutos}m"
            else:
                mesa.pedido_actual = None
                mesa.tiempo_ocupada = "-"
        else:
            mesa.pedido_actual = None
            mesa.tiempo_ocupada = "-"
    
    stats = {
        'total_mesas': total_mesas,
        'disponibles': disponibles,
        'ocupadas': ocupadas,
        'reservadas': reservadas
    }
    
    return render_template('mesas/index.html', mesas=mesas, stats=stats)

@app.route('/mesas/crear', methods=['POST'])
@login_required
@requiere_permiso('admin')
def crear_mesa():
    """Crear nueva mesa"""
    try:
        data = request.get_json()
        
        # Verificar que el número no exista
        if Mesa.query.filter_by(numero=data['numero']).first():
            return jsonify({'success': False, 'message': 'Ya existe una mesa con ese número'})
        
        mesa = Mesa(
            numero=data['numero'],
            capacidad=int(data['capacidad']),
            ubicacion=data['ubicacion'],
            descripcion=data.get('descripcion', ''),
            estado='disponible'
        )
        
        db.session.add(mesa)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Mesa creada exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/mesas/<int:id>/estado', methods=['POST'])
@login_required
def cambiar_estado_mesa(id):
    """Cambiar estado de mesa"""
    try:
        data = request.get_json()
        mesa = Mesa.query.get_or_404(id)
        
        nuevo_estado = data['estado']
        if nuevo_estado not in ['disponible', 'ocupada', 'reservada']:
            return jsonify({'success': False, 'message': 'Estado inválido'})
        
        mesa.estado = nuevo_estado
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Mesa {mesa.numero} marcada como {nuevo_estado}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/mesas/<int:id>/eliminar', methods=['DELETE'])
@login_required
@requiere_permiso('admin')
def eliminar_mesa(id):
    """Eliminar mesa"""
    try:
        mesa = Mesa.query.get_or_404(id)
        
        # Verificar que no tenga pedidos asociados
        if mesa.pedidos:
            return jsonify({'success': False, 'message': 'No se puede eliminar una mesa con pedidos asociados'})
        
        db.session.delete(mesa)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Mesa eliminada exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/mesas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@requiere_permiso('admin')
def editar_mesa(id):
    """Editar información de una mesa"""
    mesa = Mesa.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            nuevo_numero = request.form.get('numero')
            nueva_capacidad = int(request.form.get('capacidad'))
            nueva_ubicacion = request.form.get('ubicacion')
            
            # Validar que el número no esté en uso por otra mesa
            if nuevo_numero != mesa.numero:
                mesa_existente = Mesa.query.filter_by(numero=nuevo_numero, activa=True).first()
                if mesa_existente:
                    return jsonify({'success': False, 'message': f'Ya existe una mesa con el número {nuevo_numero}'})
            
            # Actualizar datos
            mesa.numero = nuevo_numero
            mesa.capacidad = nueva_capacidad
            mesa.ubicacion = nueva_ubicacion
            
            db.session.commit()
            
            return jsonify({'success': True, 'message': f'Mesa {mesa.numero} actualizada correctamente'})
            
        except ValueError:
            return jsonify({'success': False, 'message': 'La capacidad debe ser un número válido'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'Error al actualizar mesa: {str(e)}'})
    
    # GET: devolver datos de la mesa para el modal
    return jsonify({
        'id': mesa.id,
        'numero': mesa.numero,
        'capacidad': mesa.capacidad,
        'ubicacion': mesa.ubicacion or '',
        'estado': mesa.estado
    })

# ===== RUTAS DE REPORTES =====

@app.route('/reportes')
@login_required
@requiere_permiso('admin')
def reportes():
    """Dashboard de reportes con análisis completo"""
    # Período por defecto: último mes
    fecha_inicio = request.args.get('fecha_inicio', date.today().replace(day=1))
    fecha_fin = request.args.get('fecha_fin', date.today())
    
    # Convertir strings a fechas si es necesario
    if isinstance(fecha_inicio, str):
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if isinstance(fecha_fin, str):
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Consultas básicas
    pedidos_periodo = Pedido.query.filter(
        Pedido.fecha >= fecha_inicio,
        Pedido.fecha <= fecha_fin
    ).all()
    
    # Métricas principales
    ventas_totales = sum(p.total for p in pedidos_periodo)
    total_pedidos = len(pedidos_periodo)
    ticket_promedio = ventas_totales / total_pedidos if total_pedidos > 0 else 0
    productos_vendidos = sum(len(p.detalles) for p in pedidos_periodo)
    productos_unicos = len(set(d.producto_id for p in pedidos_periodo for d in p.detalles))
    
    # Calcular variación (simulada - en producción sería vs período anterior)
    variacion_ventas = 5.2  # Porcentaje simulado
    promedio_diario = total_pedidos / ((fecha_fin - fecha_inicio).days + 1)
    
    metricas = {
        'ventas_totales': ventas_totales,
        'total_pedidos': total_pedidos,
        'ticket_promedio': ticket_promedio,
        'productos_vendidos': productos_vendidos,
        'productos_unicos': productos_unicos,
        'variacion_ventas': variacion_ventas,
        'promedio_diario': round(promedio_diario, 1)
    }
    
    # Datos para gráficos
    # Ventas diarias
    ventas_diarias = {'labels': [], 'data': []}
    pedidos_por_dia = {}
    for pedido in pedidos_periodo:
        dia = pedido.fecha.strftime('%d/%m')
        if dia not in pedidos_por_dia:
            pedidos_por_dia[dia] = 0
        pedidos_por_dia[dia] += pedido.total
    
    for dia in sorted(pedidos_por_dia.keys()):
        ventas_diarias['labels'].append(dia)
        ventas_diarias['data'].append(pedidos_por_dia[dia])
    
    # Estados de pedidos
    estados = {'pendiente': 0, 'preparando': 0, 'listo': 0, 'entregado': 0}
    for pedido in pedidos_periodo:
        if pedido.estado in estados:
            estados[pedido.estado] += 1
    
    estados_pedidos = {
        'labels': ['Pendiente', 'Preparando', 'Listo', 'Entregado'],
        'data': [estados['pendiente'], estados['preparando'], estados['listo'], estados['entregado']]
    }
    
    # Ventas por hora (simulado)
    ventasHorarios = {
        'labels': ['08:00', '10:00', '12:00', '14:00', '16:00', '18:00', '20:00', '22:00'],
        'data': [150, 300, 800, 1200, 600, 900, 1100, 700]
    }
    
    # Uso de mesas (simulado)
    usoMesas = {
        'labels': ['Mesa 1', 'Mesa 2', 'Mesa 3', 'Mesa 4', 'Mesa 5'],
        'data': [12, 8, 15, 6, 10]
    }
    
    # Top productos
    from sqlalchemy import func
    try:
        top_productos = db.session.query(
            Producto.nombre,
            func.sum(DetallePedido.cantidad).label('cantidad_vendida'),
            func.sum(DetallePedido.subtotal).label('ingresos_totales')
        ).join(DetallePedido).join(Pedido).filter(
            Pedido.fecha >= fecha_inicio,
            Pedido.fecha <= fecha_fin
        ).group_by(Producto.id).order_by(
            func.sum(DetallePedido.cantidad).desc()
        ).limit(5).all()
    except:
        top_productos = []
    
    # Rendimiento por mesero
    try:
        rendimiento_meseros = db.session.query(
            Usuario.nombre_completo.label('nombre'),
            func.count(Pedido.id).label('total_pedidos'),
            func.sum(Pedido.total).label('ventas_totales'),
            func.avg(Pedido.total).label('ticket_promedio')
        ).join(Pedido, Usuario.id == Pedido.usuario_id).filter(
            Pedido.fecha >= fecha_inicio,
            Pedido.fecha <= fecha_fin,
            Usuario.rol == 'mesero'
        ).group_by(Usuario.id).order_by(
            func.sum(Pedido.total).desc()
        ).limit(5).all()
    except:
        rendimiento_meseros = []
    
    # Pedidos detallados (últimos 20)
    pedidos_detallados = Pedido.query.filter(
        Pedido.fecha >= fecha_inicio,
        Pedido.fecha <= fecha_fin
    ).order_by(Pedido.fecha.desc()).limit(20).all()
    
    # Fechas por defecto para el template
    fecha_default = {
        'inicio': fecha_inicio.strftime('%Y-%m-%d'),
        'fin': fecha_fin.strftime('%Y-%m-%d')
    }
    
    # Importar json para serializar datos
    import json
    
    return render_template('reportes/index.html',
                         metricas=metricas,
                         ventas_diarias=json.dumps(ventas_diarias),
                         estados_pedidos=json.dumps(estados_pedidos),
                         ventas_horarios=json.dumps(ventasHorarios),
                         uso_mesas=json.dumps(usoMesas),
                         top_productos=top_productos,
                         rendimiento_meseros=rendimiento_meseros,
                         pedidos_detallados=pedidos_detallados,
                         fecha_default=fecha_default)

@app.route('/reportes/exportar/<formato>')
@login_required
@requiere_permiso('admin')
def exportar_reporte(formato):
    """Exportar reportes en Excel o PDF"""
    try:
        if formato.lower() == 'excel':
            return generar_reporte_excel()
        elif formato.lower() == 'pdf':
            return generar_reporte_pdf()
        else:
            flash('Formato no válido', 'error')
            return redirect(url_for('reportes'))
    except Exception as e:
        flash(f'Error al generar reporte: {str(e)}', 'error')
        return redirect(url_for('reportes'))

def generar_reporte_excel():
    """Generar reporte en Excel con múltiples hojas"""
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Hoja 1: Resumen General
    ws1 = wb.active
    ws1.title = "Resumen General"
    
    # Headers
    headers = ["Métrica", "Valor"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Datos generales
    metricas = [
        ("Total de Pedidos", Pedido.query.count()),
        ("Total de Productos", Producto.query.count()),
        ("Total de Usuarios", Usuario.query.filter_by(activo=True).count()),
        ("Total de Mesas", Mesa.query.filter_by(activa=True).count()),
        ("Pedidos Pendientes", Pedido.query.filter_by(estado='pendiente').count()),
        ("Pedidos Completados", Pedido.query.filter_by(estado='entregado').count()),
    ]
    
    for row, (metrica, valor) in enumerate(metricas, 2):
        ws1.cell(row=row, column=1, value=metrica)
        ws1.cell(row=row, column=2, value=valor)
    
    # Hoja 2: Pedidos Detallados
    ws2 = wb.create_sheet("Pedidos")
    pedidos_headers = ["ID", "Cliente", "Mesa", "Estado", "Total", "Fecha", "Usuario"]
    
    for col, header in enumerate(pedidos_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    pedidos = Pedido.query.order_by(Pedido.fecha.desc()).limit(100).all()
    for row, pedido in enumerate(pedidos, 2):
        ws2.cell(row=row, column=1, value=pedido.id)
        ws2.cell(row=row, column=2, value=pedido.cliente_nombre)
        ws2.cell(row=row, column=3, value=pedido.mesa or "Sin mesa")
        ws2.cell(row=row, column=4, value=pedido.estado.title())
        ws2.cell(row=row, column=5, value=f"S/ {pedido.total:.2f}")
        ws2.cell(row=row, column=6, value=pedido.fecha.strftime("%Y-%m-%d %H:%M"))
        ws2.cell(row=row, column=7, value=pedido.usuario.nombre_completo if pedido.usuario else "N/A")
    
    # Hoja 3: Productos
    ws3 = wb.create_sheet("Productos")
    productos_headers = ["ID", "Nombre", "Categoría", "Precio", "Estado"]
    
    for col, header in enumerate(productos_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    productos = Producto.query.all()
    for row, producto in enumerate(productos, 2):
        ws3.cell(row=row, column=1, value=producto.id)
        ws3.cell(row=row, column=2, value=producto.nombre)
        ws3.cell(row=row, column=3, value=producto.categoria)
        ws3.cell(row=row, column=4, value=f"S/ {producto.precio:.2f}")
        ws3.cell(row=row, column=5, value="Disponible" if producto.disponible else "No disponible")
    
    # Ajustar ancho de columnas
    for ws in [ws1, ws2, ws3]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'reporte_restaurante_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def generar_reporte_pdf():
    """Generar reporte en PDF con tablas y estadísticas"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("Reporte de Restaurante", title_style))
    story.append(Spacer(1, 12))
    
    # Fecha
    story.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumen General
    story.append(Paragraph("Resumen General", styles['Heading2']))
    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total de Pedidos', str(Pedido.query.count())],
        ['Total de Productos', str(Producto.query.count())],
        ['Total de Usuarios', str(Usuario.query.filter_by(activo=True).count())],
        ['Total de Mesas', str(Mesa.query.filter_by(activa=True).count())],
        ['Pedidos Pendientes', str(Pedido.query.filter_by(estado='pendiente').count())],
        ['Pedidos Completados', str(Pedido.query.filter_by(estado='entregado').count())],
    ]
    
    resumen_table = Table(resumen_data)
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(resumen_table)
    story.append(Spacer(1, 20))
    
    # Pedidos Recientes
    story.append(Paragraph("Últimos 10 Pedidos", styles['Heading2']))
    pedidos_data = [['ID', 'Cliente', 'Mesa', 'Estado', 'Total', 'Fecha']]
    
    pedidos = Pedido.query.order_by(Pedido.fecha.desc()).limit(10).all()
    for pedido in pedidos:
        pedidos_data.append([
            str(pedido.id),
            pedido.cliente_nombre,
            pedido.mesa or "Sin mesa",
            pedido.estado.title(),
            f"S/ {pedido.total:.2f}",
            pedido.fecha.strftime('%d/%m %H:%M')
        ])
    
    pedidos_table = Table(pedidos_data)
    pedidos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(pedidos_table)
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_restaurante_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
        mimetype='application/pdf'
    )

# ===== RUTAS PRINCIPALES =====

@app.route('/')
@login_required
def index():
    """Página principal del restaurante - adaptada según rol"""
    if current_user.rol == 'cocinero':
        return redirect(url_for('dashboard_cocinero'))
    
    total_pedidos = Pedido.query.count()
    total_productos = Producto.query.count()
    total_usuarios = Usuario.query.filter_by(activo=True).count()
    total_mesas = Mesa.query.filter_by(activa=True).count()
    
    # Estadísticas del día actual
    hoy = date.today()
    pedidos_hoy = Pedido.query.filter(
        Pedido.fecha >= hoy,
        Pedido.fecha < hoy.replace(day=hoy.day + 1) if hoy.day < 28 else hoy.replace(month=hoy.month + 1, day=1)
    ).count()
    
    # Pedidos recientes según el rol
    if current_user.rol == 'mesero':
        pedidos_recientes = Pedido.query.filter_by(usuario_id=current_user.id).order_by(Pedido.fecha.desc()).limit(5).all()
    else:
        pedidos_recientes = Pedido.query.order_by(Pedido.fecha.desc()).limit(5).all()
    
    return render_template('index.html', 
                         total_pedidos=total_pedidos,
                         total_productos=total_productos,
                         total_usuarios=total_usuarios,
                         total_mesas=total_mesas,
                         pedidos_hoy=pedidos_hoy,
                         pedidos_recientes=pedidos_recientes)

@app.route('/dashboard/cocinero')
@login_required
def dashboard_cocinero():
    """Dashboard específico para cocineros"""
    if current_user.rol != 'cocinero':
        flash('Acceso denegado', 'error')
        return redirect(url_for('index'))
    
    # Estadísticas para cocineros
    pendientes = Pedido.query.filter_by(estado='pendiente').count()
    preparando = Pedido.query.filter_by(estado='preparando').count()
    listos = Pedido.query.filter_by(estado='listo').count()
    
    # Pedidos pendientes más antiguos (prioritarios)
    pedidos_pendientes = Pedido.query.filter_by(estado='pendiente').order_by(Pedido.fecha.asc()).limit(10).all()
    
    # Pedidos en preparación asignados al cocinero actual
    mis_preparando = Pedido.query.filter_by(estado='preparando', cocinero_id=current_user.id).order_by(Pedido.fecha.asc()).all()
    
    # Pedidos listos para servir
    pedidos_listos = Pedido.query.filter_by(estado='listo').order_by(Pedido.fecha.desc()).limit(5).all()
    
    return render_template('dashboard_cocinero.html',
                         pendientes=pendientes,
                         preparando=preparando,
                         listos=listos,
                         pedidos_pendientes=pedidos_pendientes,
                         mis_preparando=mis_preparando,
                         pedidos_listos=pedidos_listos)

@app.route('/menu')
def menu():
    """Mostrar el menú completo (público)"""
    productos = Producto.query.filter_by(disponible=True).all()
    return render_template('menu.html', productos=productos)

# ===== RUTAS DE PEDIDOS =====

@app.route('/pedidos/')
@login_required
def lista_pedidos():
    """Mostrar lista de pedidos según el rol del usuario"""
    page = request.args.get('page', 1, type=int)
    estado = request.args.get('estado', 'todos')
    
    # Query base
    query = Pedido.query
    
    # Filtrar según el rol
    if current_user.rol == 'mesero':
        # Los meseros solo ven sus propios pedidos
        query = query.filter_by(usuario_id=current_user.id)
    elif current_user.rol == 'cocinero':
        # Los cocineros ven solo pedidos en estados relevantes para cocina
        query = query.filter(Pedido.estado.in_(['pendiente', 'preparando', 'listo']))
    # Los admin ven todos los pedidos (sin filtro adicional)
    
    # Aplicar filtro de estado
    if estado != 'todos':
        query = query.filter_by(estado=estado)
    
    pedidos = query.order_by(Pedido.fecha.desc()).paginate(
        page=page, per_page=10, error_out=False)
    
    return render_template('pedidos/lista.html', pedidos=pedidos, estado_filtro=estado)

@app.route('/pedidos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_pedido():
    """Crear un nuevo pedido"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cliente_nombre = request.form.get('cliente_nombre')
            cliente_telefono = request.form.get('cliente_telefono')
            mesa_numero = request.form.get('mesa')
            observaciones = request.form.get('observaciones')
            
            # Buscar la mesa por número
            mesa_obj = None
            if mesa_numero:
                mesa_obj = Mesa.query.filter_by(numero=mesa_numero, activa=True).first()
                if mesa_obj and mesa_obj.estado == 'ocupada':
                    flash(f'La mesa {mesa_numero} ya está ocupada', 'error')
                    productos = Producto.query.filter_by(disponible=True).order_by(Producto.categoria, Producto.nombre).all()
                    productos_dict = [producto.to_dict() for producto in productos]
                    mesas = Mesa.query.filter_by(activa=True).all()
                    return render_template('pedidos/nuevo.html', productos=productos, productos_json=productos_dict, mesas=mesas)
            
            # Crear el pedido
            pedido = Pedido(
                cliente_nombre=cliente_nombre,
                cliente_telefono=cliente_telefono,
                mesa_id=mesa_obj.id if mesa_obj else None,
                mesa_numero=mesa_numero,  # Mantener compatibilidad
                observaciones=observaciones,
                estado='pendiente',
                usuario_id=current_user.id
            )
            
            db.session.add(pedido)
            db.session.flush()  # Para obtener el ID del pedido
            
            # Cambiar estado de la mesa a ocupada si existe
            if mesa_obj:
                mesa_obj.estado = 'ocupada'
            
            # Procesar productos seleccionados
            productos_ids = request.form.getlist('producto_id')
            cantidades = request.form.getlist('cantidad')
            observaciones_detalle = request.form.getlist('observaciones_detalle')
            
            total_pedido = 0
            
            for i, producto_id in enumerate(productos_ids):
                if producto_id and cantidades[i]:
                    producto = db.session.get(Producto, int(producto_id))
                    cantidad = int(cantidades[i])
                    
                    if producto and cantidad > 0:
                        detalle = DetallePedido(
                            pedido_id=pedido.id,
                            producto_id=producto.id,
                            cantidad=cantidad,
                            precio_unitario=producto.precio,
                            observaciones=observaciones_detalle[i] if i < len(observaciones_detalle) else None
                        )
                        detalle.calcular_subtotal()
                        db.session.add(detalle)
                        total_pedido += detalle.subtotal
            
            pedido.total = total_pedido
            db.session.commit()
            
            flash(f'Pedido #{pedido.id} creado exitosamente. Mesa {mesa_numero} marcada como ocupada.', 'success')
            return redirect(url_for('ver_pedido', id=pedido.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el pedido: {str(e)}', 'error')
    
    # GET: mostrar formulario
    productos = Producto.query.filter_by(disponible=True).order_by(Producto.categoria, Producto.nombre).all()
    # Convertir productos a diccionarios para JSON
    productos_dict = [producto.to_dict() for producto in productos]
    # Obtener mesas disponibles
    mesas = Mesa.query.filter_by(activa=True).all()
    return render_template('pedidos/nuevo.html', productos=productos, productos_json=productos_dict, mesas=mesas)

@app.route('/pedidos/<int:id>')
def ver_pedido(id):
    """Ver detalles de un pedido específico"""
    pedido = Pedido.query.get_or_404(id)
    return render_template('pedidos/detalle.html', pedido=pedido)

@app.route('/pedidos/<int:id>/cambiar_estado', methods=['POST'])
def cambiar_estado(id):
    """Cambiar el estado de un pedido"""
    try:
        pedido = Pedido.query.get_or_404(id)
        nuevo_estado = request.form.get('estado')
        
        estados_validos = ['pendiente', 'preparando', 'listo', 'entregado', 'cancelado']
        
        if nuevo_estado in estados_validos:
            estado_anterior = pedido.estado
            pedido.estado = nuevo_estado
            
            # Si un cocinero toma un pedido (pasa a preparando), asignarlo
            if nuevo_estado == 'preparando' and current_user.rol == 'cocinero':
                pedido.cocinero_id = current_user.id
            
            # Si el pedido se entrega o cancela, liberar la mesa
            if nuevo_estado in ['entregado', 'cancelado'] and pedido.mesa_info:
                pedido.mesa_info.estado = 'disponible'
                flash(f'Estado del pedido #{pedido.id} cambiado a {nuevo_estado}. Mesa {pedido.mesa_info.numero} liberada.', 'success')
            else:
                flash(f'Estado del pedido #{pedido.id} cambiado a {nuevo_estado}', 'success')
                
            db.session.commit()
            
            # Si es una petición AJAX (desde dashboard cocinero), devolver JSON
            if request.headers.get('Content-Type') == 'application/json' or request.is_json:
                return jsonify({'success': True, 'message': f'Estado cambiado a {nuevo_estado}'})
                
        else:
            flash('Estado no válido', 'error')
            if request.headers.get('Content-Type') == 'application/json' or request.is_json:
                return jsonify({'success': False, 'message': 'Estado no válido'})
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar estado: {str(e)}', 'error')
        if request.headers.get('Content-Type') == 'application/json' or request.is_json:
            return jsonify({'success': False, 'message': str(e)})
    
    return redirect(url_for('ver_pedido', id=id))

@app.route('/pedidos/<int:id>/eliminar', methods=['POST'])
def eliminar_pedido(id):
    """Eliminar un pedido"""
    pedido = Pedido.query.get_or_404(id)
    
    try:
        # Si el pedido tenía una mesa asignada, liberarla
        if pedido.mesa_info:
            pedido.mesa_info.estado = 'disponible'
            flash(f'Pedido #{id} eliminado correctamente. Mesa {pedido.mesa_info.numero} liberada.', 'success')
        else:
            flash(f'Pedido #{id} eliminado correctamente', 'success')
            
        db.session.delete(pedido)
        db.session.commit()
        return redirect(url_for('lista_pedidos'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el pedido: {str(e)}', 'error')
        return redirect(url_for('ver_pedido', id=id))

# ===== RUTAS DE PRODUCTOS =====

@app.route('/productos/')
def lista_productos():
    """Mostrar lista de todos los productos"""
    categoria = request.args.get('categoria', 'todos')
    buscar = request.args.get('buscar', '')
    
    query = Producto.query
    
    if categoria != 'todos':
        query = query.filter_by(categoria=categoria)
    
    if buscar:
        query = query.filter(Producto.nombre.contains(buscar))
    
    productos = query.order_by(Producto.categoria, Producto.nombre).all()
    categorias = db.session.query(Producto.categoria).distinct().all()
    categorias = [cat[0] for cat in categorias]
    
    return render_template('productos/lista.html', 
                         productos=productos, 
                         categorias=categorias,
                         categoria_filtro=categoria,
                         buscar=buscar)

@app.route('/productos/nuevo', methods=['GET', 'POST'])
def nuevo_producto():
    """Crear un nuevo producto"""
    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre')
            descripcion = request.form.get('descripcion')
            precio = float(request.form.get('precio'))
            categoria = request.form.get('categoria')
            nueva_categoria = request.form.get('nueva_categoria')
            disponible = 'disponible' in request.form
            
            # Si se especifica una nueva categoría, usarla
            if nueva_categoria:
                categoria = nueva_categoria
            
            producto = Producto(
                nombre=nombre,
                descripcion=descripcion,
                precio=precio,
                categoria=categoria,
                disponible=disponible
            )
            
            db.session.add(producto)
            db.session.commit()
            
            flash(f'Producto "{nombre}" creado exitosamente', 'success')
            return redirect(url_for('lista_productos'))
            
        except ValueError:
            flash('Por favor ingresa un precio válido', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el producto: {str(e)}', 'error')
    
    # Obtener categorías existentes para el formulario
    categorias = db.session.query(Producto.categoria).distinct().all()
    categorias = [cat[0] for cat in categorias]
    
    return render_template('productos/nuevo.html', categorias=categorias)

@app.route('/productos/<int:id>')
def ver_producto(id):
    """Ver detalles de un producto específico"""
    producto = Producto.query.get_or_404(id)
    return render_template('productos/detalle.html', producto=producto)

@app.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
def editar_producto(id):
    """Editar un producto existente"""
    producto = Producto.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            producto.nombre = request.form.get('nombre')
            producto.descripcion = request.form.get('descripcion')
            producto.precio = float(request.form.get('precio'))
            
            categoria = request.form.get('categoria')
            nueva_categoria = request.form.get('nueva_categoria')
            
            # Si se especifica una nueva categoría, usarla
            if nueva_categoria:
                producto.categoria = nueva_categoria
            else:
                producto.categoria = categoria
                
            producto.disponible = 'disponible' in request.form
            
            db.session.commit()
            
            flash(f'Producto "{producto.nombre}" actualizado exitosamente', 'success')
            return redirect(url_for('ver_producto', id=id))
            
        except ValueError:
            flash('Por favor ingresa un precio válido', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {str(e)}', 'error')
    
    # Obtener categorías existentes para el formulario
    categorias = db.session.query(Producto.categoria).distinct().all()
    categorias = [cat[0] for cat in categorias]
    
    return render_template('productos/editar.html', producto=producto, categorias=categorias)

@app.route('/productos/<int:id>/toggle_disponibilidad', methods=['POST'])
def toggle_disponibilidad(id):
    """Cambiar la disponibilidad de un producto"""
    producto = Producto.query.get_or_404(id)
    producto.disponible = not producto.disponible
    
    try:
        db.session.commit()
        estado = "disponible" if producto.disponible else "no disponible"
        flash(f'Producto "{producto.nombre}" marcado como {estado}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar disponibilidad: {str(e)}', 'error')
    
    return redirect(url_for('ver_producto', id=id))

@app.route('/productos/<int:id>/eliminar', methods=['POST'])
def eliminar_producto(id):
    """Eliminar un producto"""
    producto = Producto.query.get_or_404(id)
    
    try:
        # Verificar si el producto tiene pedidos asociados
        if producto.detalles_pedido:
            flash(f'No se puede eliminar "{producto.nombre}" porque tiene pedidos asociados. '
                  'Puedes marcarlo como no disponible en su lugar.', 'warning')
            return redirect(url_for('ver_producto', id=id))
        
        nombre = producto.nombre
        db.session.delete(producto)
        db.session.commit()
        
        flash(f'Producto "{nombre}" eliminado correctamente', 'success')
        return redirect(url_for('lista_productos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el producto: {str(e)}', 'error')
        return redirect(url_for('ver_producto', id=id))

def crear_datos_iniciales():
    """Crear datos iniciales para el restaurante"""
    
    # Crear usuario administrador por defecto
    if Usuario.query.count() == 0:
        admin = Usuario(
            username='admin',
            email='admin@restaurante.com',
            nombre_completo='Administrador del Sistema',
            rol='admin'
        )
        admin.set_password('admin123')  # Cambiar en producción
        db.session.add(admin)
        
        # Crear usuarios de ejemplo
        mesero1 = Usuario(
            username='mesero1',
            email='mesero1@restaurante.com',
            nombre_completo='Juan Pérez',
            rol='mesero'
        )
        mesero1.set_password('mesero123')
        db.session.add(mesero1)
        
        cocinero1 = Usuario(
            username='cocinero1',
            email='cocinero1@restaurante.com',
            nombre_completo='María González',
            rol='cocinero'
        )
        cocinero1.set_password('cocinero123')
        db.session.add(cocinero1)
        
        print("Usuarios iniciales creados")
    
    # Crear mesas por defecto
    if Mesa.query.count() == 0:
        mesas_iniciales = [
            Mesa(numero='1', capacidad=2, ubicacion='interior'),
            Mesa(numero='2', capacidad=4, ubicacion='interior'),
            Mesa(numero='3', capacidad=4, ubicacion='interior'),
            Mesa(numero='4', capacidad=6, ubicacion='interior'),
            Mesa(numero='5', capacidad=2, ubicacion='terraza'),
            Mesa(numero='6', capacidad=4, ubicacion='terraza'),
            Mesa(numero='VIP1', capacidad=8, ubicacion='vip'),
        ]
        
        for mesa in mesas_iniciales:
            db.session.add(mesa)
        
        print("Mesas iniciales creadas")
    
    # Crear productos iniciales
    if Producto.query.count() == 0:
        productos_iniciales = [
            Producto(nombre='Hamburguesa Clásica', precio=18.50, categoria='Principal', disponible=True,
                    descripcion='Hamburguesa de carne con lechuga, tomate, cebolla y papas fritas'),
            Producto(nombre='Pizza Margherita', precio=25.00, categoria='Principal', disponible=True,
                    descripcion='Pizza tradicional con tomate, mozzarella y albahaca fresca'),
            Producto(nombre='Ensalada César', precio=15.00, categoria='Ensalada', disponible=True,
                    descripcion='Lechuga romana, crutones, queso parmesano y aderezo césar'),
            Producto(nombre='Papas Fritas', precio=8.00, categoria='Acompañamiento', disponible=True,
                    descripcion='Papas fritas doradas y crujientes con sal marina'),
            Producto(nombre='Limonada', precio=6.00, categoria='Bebida', disponible=True,
                    descripcion='Limonada natural refrescante con hielo y menta'),
            Producto(nombre='Pollo a la Plancha', precio=22.00, categoria='Principal', disponible=True,
                    descripcion='Pechuga de pollo a la plancha con verduras y arroz'),
            Producto(nombre='Ceviche Mixto', precio=28.00, categoria='Marina', disponible=True,
                    descripcion='Ceviche de pescado y mariscos con camote y choclo'),
            Producto(nombre='Lomo Saltado', precio=26.00, categoria='Principal', disponible=True,
                    descripcion='Lomo de res saltado con cebolla, tomate y papas fritas'),
        ]
        
        for producto in productos_iniciales:
            db.session.add(producto)
        
        print("Productos iniciales creados")
    
    # Crear pedidos de ejemplo
    if Pedido.query.count() == 0:
        # Obtener usuarios y mesa para el ejemplo
        admin = Usuario.query.filter_by(username='admin').first()
        mesero = Usuario.query.filter_by(rol='mesero').first()
        mesa1 = Mesa.query.filter_by(numero='1').first()
        mesa2 = Mesa.query.filter_by(numero='2').first()
        
        if admin and mesa1:
            # Pedido ejemplo 1
            pedido1 = Pedido(
                cliente_nombre='Carlos Mendoza',
                cliente_telefono='987654321',
                mesa_id=mesa1.id,
                mesa_numero=mesa1.numero,
                estado='entregado',
                usuario_id=mesero.id if mesero else admin.id,
                observaciones='Sin cebolla'
            )
            
            # Pedido ejemplo 2
            pedido2 = Pedido(
                cliente_nombre='Ana Torres',
                cliente_telefono='123456789',
                mesa_id=mesa2.id if mesa2 else mesa1.id,
                mesa_numero=mesa2.numero if mesa2 else mesa1.numero,
                estado='preparando',
                usuario_id=admin.id,
                observaciones='Extra salsa'
            )
            
            db.session.add(pedido1)
            db.session.add(pedido2)
            
            # Commit para obtener IDs
            db.session.commit()
            
            # Productos para los detalles
            hamburguesa = Producto.query.filter_by(nombre='Hamburguesa Clásica').first()
            papas = Producto.query.filter_by(nombre='Papas Fritas').first()
            limonada = Producto.query.filter_by(nombre='Limonada').first()
            
            if hamburguesa and papas and limonada:
                # Detalles pedido 1
                detalle1 = DetallePedido(
                    pedido_id=pedido1.id,
                    producto_id=hamburguesa.id,
                    cantidad=1,
                    precio_unitario=hamburguesa.precio
                )
                detalle1.calcular_subtotal()
                
                detalle2 = DetallePedido(
                    pedido_id=pedido1.id,
                    producto_id=papas.id,
                    cantidad=1,
                    precio_unitario=papas.precio
                )
                detalle2.calcular_subtotal()
                
                # Detalles pedido 2
                detalle3 = DetallePedido(
                    pedido_id=pedido2.id,
                    producto_id=limonada.id,
                    cantidad=2,
                    precio_unitario=limonada.precio
                )
                detalle3.calcular_subtotal()
                
                db.session.add_all([detalle1, detalle2, detalle3])
                
                # Calcular totales
                pedido1.calcular_total()
                pedido2.calcular_total()
        
        print("Pedidos de ejemplo creados")
    
    db.session.commit()
    print("Datos iniciales creados correctamente")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        crear_datos_iniciales()
    
    # Configuración para despliegue en producción
    import os
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    
    if debug:
        print("Servidor iniciado en http://127.0.0.1:5000")
        app.run(host='127.0.0.1', port=5000, debug=True)
    else:
        print(f"Servidor iniciado en producción en puerto {port}")
        app.run(host='0.0.0.0', port=port, debug=False)
        